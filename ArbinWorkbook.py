import os
import openpyxl

# =============================================================================
# ArbinWorkbook
# -----------------------------------------------------------------------------
# 
# 
# 
# 
#
# =============================================================================

class ArbinWorkbook( object ):

    def __init__( self, file_name ):
        self.file_name = file_name + ".xlsx"
        
        self.wb = openpyxl.Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Global Info"
        self.ws2 = self.wb.create_sheet("Channel")
        self.ws3 = self.wb.create_sheet("Statistics")


    def save_workbook( self, path ):
        if( not os.path.isdir(path) ): os.makedirs(path)
        self.wb.save( path + self.file_name )
    
  
    
    # --------------------------------------------------------------------------------------
    # Utilities
    # --------------------------------------------------------------------------------------
    @staticmethod
    def background_color( worksheet, row, color ):
        backgroundFill = openpyxl.styles.fills.PatternFill( start_color=color, end_color=color, fill_type='solid' )
        for cell in list(worksheet.rows)[row]:
            cell.fill = backgroundFill
            
            
    @staticmethod
    def bottom_border( worksheet, row ):
        border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thin'))       
        for cell in list(worksheet.rows)[row]:
            cell.border = border


    @staticmethod
    def right_justify( worksheet, column_row ):
        alignment = openpyxl.styles.Alignment(horizontal='right')
        cell = worksheet[column_row]
        cell.alignment = alignment


    @staticmethod
    def resize_cells( worksheet, rows_to_check ):
        dims = {}
        for row in list(worksheet.rows)[rows_to_check]:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max( (dims.get(cell.column_letter, 0), len(str(cell.value))) ) 
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value
        
