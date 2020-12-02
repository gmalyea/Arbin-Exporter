import os
import openpyxl

# =============================================================================
# ArbinWorkbook
# -----------------------------------------------------------------------------
# 
# 
# 
# 
#
# =============================================================================


class ArbinWorkbook( object ):

    # Class Initialization
    # -----------------------------------------------------------------------------
    def __init__( self, file_name ):
        self.file_name = file_name + ".xlsx"
        
        self.wb = openpyxl.Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Global Info"
        self.ws2 = self.wb.create_sheet("Channel")
        self.ws3 = self.wb.create_sheet("Statistics By Cycle")

    def save_workbook( self, path ):
        path_norm = os.path.normpath( path )
        file_norm = os.path.normpath( self.file_name )
        
        if( not os.path.isdir(path_norm) ):
            os.makedirs(path_norm)
        self.wb.save( os.path.join(path_norm, file_norm) )
        
    
    
    # --------------------------------------------------------------------------------------
    # Utilities
    # --------------------------------------------------------------------------------------
    @staticmethod
    def background_color( worksheet, row, color ):
        for cell in list(worksheet.rows)[row]:
            cell.fill = openpyxl.styles.fills.PatternFill( start_color=color, end_color=color, fill_type='solid' )
            
            
    @staticmethod
    def border_bottom( worksheet, row ):
        for cell in list(worksheet.rows)[row]:
            cell.border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thin'))     


    @staticmethod
    def justify_right( worksheet, column_row ):
        cell = worksheet[column_row]
        cell.alignment = openpyxl.styles.Alignment(horizontal='right')


    @staticmethod
    def justify_left( worksheet, column_row ):
        cell = worksheet[column_row]
        cell.alignment = openpyxl.styles.Alignment(horizontal='left')
        

    @staticmethod
    def resize_cells( worksheet, rows_to_check ):
        dims = {}
        for row in list(worksheet.rows)[rows_to_check]:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max( (dims.get(cell.column_letter, 0), len(str(cell.value))) ) 
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value
